"""
BaseIOService - Import/Export Générique
Élimine duplication entre tous les modules

Formats supportés:
- CSV (lecture/écriture)
- JSON (lecture/écriture)
- Excel (lecture/écriture avec openpyxl)
- Markdown (écriture uniquement)
"""
import csv
import json
import io
import logging
from typing import List, Dict, Optional, Any, Type, Callable
from datetime import datetime, date
from dataclasses import dataclass

from src.core.errors import handle_errors, ValidationError

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════

@dataclass
class IOConfig:
    """Configuration import/export"""

    # Champs
    field_mapping: Dict[str, str]  # {"db_field": "export_label"}
    required_fields: List[str]     # Champs obligatoires import

    # Formats
    date_format: str = "%d/%m/%Y"
    datetime_format: str = "%d/%m/%Y %H:%M"

    # Validation
    validator: Optional[Callable] = None  # Fonction validation custom

    # Transformation
    pre_export: Optional[Callable] = None  # Transform avant export
    post_import: Optional[Callable] = None  # Transform après import

    # CSV
    csv_delimiter: str = ","
    csv_encoding: str = "utf-8"

    # Excel
    excel_sheet_name: str = "Data"


# ═══════════════════════════════════════════════════════════════
# BASE IO SERVICE
# ═══════════════════════════════════════════════════════════════

class BaseIOService:
    """
    Service I/O générique pour tous les modules

    Usage:
        config = IOConfig(...)
        io_service = BaseIOService(config)

        # Export
        csv_str = io_service.to_csv(items)

        # Import
        items, errors = io_service.from_csv(csv_str)
    """

    def __init__(self, config: IOConfig):
        self.config = config

    # ═══════════════════════════════════════════════════════════
    # EXPORT CSV
    # ═══════════════════════════════════════════════════════════

    @handle_errors(show_in_ui=False, fallback_value="")
    def to_csv(self, items: List[Any]) -> str:
        """
        Exporte en CSV

        Args:
            items: Liste d'objets ou dicts

        Returns:
            String CSV
        """
        if not items:
            return ""

        # Transformer si nécessaire
        if self.config.pre_export:
            items = [self.config.pre_export(item) for item in items]

        # Convertir en dicts si objets
        data = [self._to_dict(item) for item in items]

        # Créer CSV
        output = io.StringIO()

        # Headers (labels d'export)
        fieldnames = list(self.config.field_mapping.values())

        writer = csv.DictWriter(
            output,
            fieldnames=fieldnames,
            delimiter=self.config.csv_delimiter
        )
        writer.writeheader()

        # Lignes
        for item_dict in data:
            row = {}
            for db_field, export_label in self.config.field_mapping.items():
                value = item_dict.get(db_field)
                row[export_label] = self._format_value(value)

            writer.writerow(row)

        logger.info(f"CSV export: {len(items)} items")
        return output.getvalue()

    # ═══════════════════════════════════════════════════════════
    # IMPORT CSV
    # ═══════════════════════════════════════════════════════════

    @handle_errors(show_in_ui=False, fallback_value=([], []))
    def from_csv(self, csv_str: str) -> tuple[List[Dict], List[str]]:
        """
        Importe depuis CSV

        Args:
            csv_str: Contenu CSV

        Returns:
            (items_validés, erreurs)
        """
        reader = csv.DictReader(
            io.StringIO(csv_str),
            delimiter=self.config.csv_delimiter
        )

        items = []
        errors = []

        # Mapping inverse (label → db_field)
        inverse_mapping = {
            v: k for k, v in self.config.field_mapping.items()
        }

        for row_num, row in enumerate(reader, start=2):
            try:
                # Convertir labels → db fields
                item_dict = {}
                for export_label, value in row.items():
                    if export_label in inverse_mapping:
                        db_field = inverse_mapping[export_label]
                        item_dict[db_field] = self._parse_value(value)

                # Validation champs requis
                missing = [
                    f for f in self.config.required_fields
                    if not item_dict.get(f)
                ]

                if missing:
                    errors.append(f"Ligne {row_num}: Champs manquants: {missing}")
                    continue

                # Validation custom
                if self.config.validator:
                    valid, error = self.config.validator(item_dict)
                    if not valid:
                        errors.append(f"Ligne {row_num}: {error}")
                        continue

                # Transformation post-import
                if self.config.post_import:
                    item_dict = self.config.post_import(item_dict)

                items.append(item_dict)

            except Exception as e:
                errors.append(f"Ligne {row_num}: {str(e)}")

        logger.info(f"CSV import: {len(items)} items, {len(errors)} errors")
        return items, errors

    # ═══════════════════════════════════════════════════════════
    # EXPORT JSON
    # ═══════════════════════════════════════════════════════════

    @handle_errors(show_in_ui=False, fallback_value="[]")
    def to_json(self, items: List[Any], indent: int = 2) -> str:
        """
        Exporte en JSON

        Args:
            items: Liste d'objets ou dicts
            indent: Indentation

        Returns:
            String JSON
        """
        if not items:
            return "[]"

        # Transformer
        if self.config.pre_export:
            items = [self.config.pre_export(item) for item in items]

        # Convertir en dicts
        data = [self._to_dict(item) for item in items]

        # Filtrer selon mapping
        filtered = []
        for item_dict in data:
            filtered_item = {
                db_field: item_dict.get(db_field)
                for db_field in self.config.field_mapping.keys()
                if db_field in item_dict
            }
            filtered.append(filtered_item)

        logger.info(f"JSON export: {len(items)} items")
        return json.dumps(filtered, indent=indent, ensure_ascii=False, default=str)

    # ═══════════════════════════════════════════════════════════
    # IMPORT JSON
    # ═══════════════════════════════════════════════════════════

    @handle_errors(show_in_ui=False, fallback_value=([], []))
    def from_json(self, json_str: str) -> tuple[List[Dict], List[str]]:
        """
        Importe depuis JSON

        Args:
            json_str: Contenu JSON

        Returns:
            (items_validés, erreurs)
        """
        try:
            data = json.loads(json_str)
        except json.JSONDecodeError as e:
            return [], [f"JSON invalide: {str(e)}"]

        if not isinstance(data, list):
            data = [data]

        items = []
        errors = []

        for idx, item_dict in enumerate(data, start=1):
            try:
                # Validation champs requis
                missing = [
                    f for f in self.config.required_fields
                    if f not in item_dict or not item_dict[f]
                ]

                if missing:
                    errors.append(f"Item {idx}: Champs manquants: {missing}")
                    continue

                # Validation custom
                if self.config.validator:
                    valid, error = self.config.validator(item_dict)
                    if not valid:
                        errors.append(f"Item {idx}: {error}")
                        continue

                # Transformation
                if self.config.post_import:
                    item_dict = self.config.post_import(item_dict)

                items.append(item_dict)

            except Exception as e:
                errors.append(f"Item {idx}: {str(e)}")

        logger.info(f"JSON import: {len(items)} items, {len(errors)} errors")
        return items, errors

    # ═══════════════════════════════════════════════════════════
    # EXPORT EXCEL
    # ═══════════════════════════════════════════════════════════

    @handle_errors(show_in_ui=False, fallback_value=None)
    def to_excel(self, items: List[Any]) -> bytes:
        """
        Exporte en Excel (nécessite openpyxl)

        Args:
            items: Liste d'objets ou dicts

        Returns:
            Bytes Excel
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl non installé")
            raise ValidationError(
                "openpyxl requis pour Excel",
                user_message="Installer: pip install openpyxl"
            )

        if not items:
            return None

        # Transformer
        if self.config.pre_export:
            items = [self.config.pre_export(item) for item in items]

        # Convertir
        data = [self._to_dict(item) for item in items]

        # Créer workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.config.excel_sheet_name

        # Headers
        headers = list(self.config.field_mapping.values())
        ws.append(headers)

        # Données
        for item_dict in data:
            row = []
            for db_field in self.config.field_mapping.keys():
                value = item_dict.get(db_field)
                row.append(self._format_value(value))
            ws.append(row)

        # Style headers (gras)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Ajuster largeurs colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Sauver en bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel export: {len(items)} items")
        return output.getvalue()

    # ═══════════════════════════════════════════════════════════
    # EXPORT MARKDOWN
    # ═══════════════════════════════════════════════════════════

    @handle_errors(show_in_ui=False, fallback_value="")
    def to_markdown(self, items: List[Any], title: str = "Export") -> str:
        """
        Exporte en Markdown

        Args:
            items: Liste d'objets ou dicts
            title: Titre du document

        Returns:
            String Markdown
        """
        if not items:
            return ""

        # Transformer
        if self.config.pre_export:
            items = [self.config.pre_export(item) for item in items]

        data = [self._to_dict(item) for item in items]

        # Construire Markdown
        md = f"# {title}\n\n"
        md += f"*Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}*\n\n"
        md += f"**Total:** {len(items)} élément(s)\n\n"
        md += "---\n\n"

        # Table
        headers = list(self.config.field_mapping.values())

        # Header row
        md += "| " + " | ".join(headers) + " |\n"
        md += "| " + " | ".join(["---"] * len(headers)) + " |\n"

        # Data rows
        for item_dict in data:
            row = []
            for db_field in self.config.field_mapping.keys():
                value = item_dict.get(db_field, "")
                row.append(str(self._format_value(value)))

            md += "| " + " | ".join(row) + " |\n"

        logger.info(f"Markdown export: {len(items)} items")
        return md

    # ═══════════════════════════════════════════════════════════
    # HELPERS
    # ═══════════════════════════════════════════════════════════

    def _to_dict(self, item: Any) -> Dict:
        """Convertit objet en dict"""
        if isinstance(item, dict):
            return item

        # SQLAlchemy model
        if hasattr(item, "__table__"):
            result = {}
            for col in item.__table__.columns:
                result[col.name] = getattr(item, col.name)
            return result

        # Dataclass/Pydantic
        if hasattr(item, "__dict__"):
            return item.__dict__

        return {}

    def _format_value(self, value: Any) -> str:
        """Formate valeur pour export"""
        if value is None:
            return ""

        if isinstance(value, (date, datetime)):
            if isinstance(value, datetime):
                return value.strftime(self.config.datetime_format)
            return value.strftime(self.config.date_format)

        if isinstance(value, bool):
            return "Oui" if value else "Non"

        if isinstance(value, (list, tuple)):
            return ", ".join(str(v) for v in value)

        return str(value)

    def _parse_value(self, value: str) -> Any:
        """Parse valeur depuis import"""
        if not value or value.strip() == "":
            return None

        value = value.strip()

        # Boolean
        if value.lower() in ["oui", "yes", "true", "1"]:
            return True
        if value.lower() in ["non", "no", "false", "0"]:
            return False

        # Number
        try:
            if "." in value or "," in value:
                return float(value.replace(",", "."))
            return int(value)
        except ValueError:
            pass

        # Date
        for fmt in [self.config.date_format, "%Y-%m-%d", "%d-%m-%Y"]:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue

        # String
        return value


# ═══════════════════════════════════════════════════════════════
# FACTORY
# ═══════════════════════════════════════════════════════════════

def create_io_service(config: IOConfig) -> BaseIOService:
    """Factory pour créer service I/O"""
    return BaseIOService(config)


# ═══════════════════════════════════════════════════════════════
# TEMPLATES
# ═══════════════════════════════════════════════════════════════

def get_csv_template(config: IOConfig) -> str:
    """Retourne template CSV vide"""
    output = io.StringIO()

    headers = list(config.field_mapping.values())
    writer = csv.DictWriter(output, fieldnames=headers, delimiter=config.csv_delimiter)
    writer.writeheader()

    return output.getvalue()